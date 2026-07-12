"""
generate_annotation_sheet.py  —  Step 11b (Human Annotation Spreadsheet)

Randomly samples 200 pairs from the claim comparison results and exports
them to an Excel workbook ready for manual annotation by two independent
raters.

The workbook contains three sheets:
  1. Instructions  — annotation guidelines for raters
  2. Annotation    — the 200 pairs to annotate (one row per pair)
  3. Codebook      — definitions of each change type

After both raters have filled in their columns, run:
    python3 code/analysis/compute_agreement.py

Reads from:
    data/claims/claim_changes.jsonl
    data/final/analysis_corpus.csv

Writes to:
    data/validation/annotation_sheet.xlsx

Usage:
    cd SSRN_bioRxiv_medRxiv_data_collection_via_Dimensions
    pip install openpyxl
    python3 code/analysis/generate_annotation_sheet.py
"""

import json
import random
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]

CHANGES_FILE = PROJECT_ROOT / "data" / "claims" / "claim_changes.jsonl"
CORPUS_FILE  = PROJECT_ROOT / "data" / "final"  / "analysis_corpus.csv"
OUT_FILE     = PROJECT_ROOT / "data" / "validation" / "annotation_sheet.xlsx"

OUT_FILE.parent.mkdir(parents=True, exist_ok=True)

SAMPLE_N   = 200
RANDOM_SEED = 42

# ── Colours ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")
YELLOW_FILL   = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")
GREY_FILL     = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT     = Font(bold=True, size=10)
NORMAL_FONT   = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

VALID_CHANGES = ["strengthened", "weakened", "unchanged", "removed", "added"]

# ── Load data ──────────────────────────────────────────────────────────────────

def load_changes():
    records = []
    with open(CHANGES_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
                if not r.get("error") and r.get("comparisons"):
                    records.append(r)
            except json.JSONDecodeError:
                pass
    return records


def load_corpus():
    df = pd.read_csv(CORPUS_FILE, dtype=str, low_memory=False)
    return df.set_index("doi") if "doi" in df.columns else df

# ── Sample ─────────────────────────────────────────────────────────────────────

def sample_records(records, n=SAMPLE_N, seed=RANDOM_SEED):
    random.seed(seed)
    # Stratify by dominant_change to get a representative mix
    by_type = {}
    for r in records:
        dc = r.get("dominant_change", "other")
        by_type.setdefault(dc, []).append(r)

    sampled = []
    per_type = max(1, n // len(by_type))
    for dc, recs in by_type.items():
        k = min(per_type, len(recs))
        sampled.extend(random.sample(recs, k))

    # Top up to exactly n
    remaining = [r for r in records if r not in sampled]
    random.shuffle(remaining)
    sampled.extend(remaining[:max(0, n - len(sampled))])
    sampled = sampled[:n]
    random.shuffle(sampled)
    return sampled


def fmt_claims(claims_list):
    """Format a list of claim dicts (or strings) as a numbered string."""
    lines = []
    for i, c in enumerate(claims_list, 1):
        if isinstance(c, dict):
            cert  = c.get("certainty", "")
            claim = c.get("claim", "")
            cert_tag = f"[{cert}] " if cert else ""
            lines.append(f"{i}. {cert_tag}{claim}")
        else:
            lines.append(f"{i}. {str(c)}")
    return "\n".join(lines)


def fmt_comparisons(comparisons):
    """Format comparison list as a readable string."""
    lines = []
    for comp in comparisons:
        if isinstance(comp, dict):
            pre   = comp.get("preprint_claim", "")
            pub   = comp.get("pub_claim", "")
            ct    = comp.get("change_type", "")
            lines.append(f"[{ct.upper()}]")
            lines.append(f"  Preprint : {pre}")
            lines.append(f"  Published: {pub}")
            lines.append("")
    return "\n".join(lines).strip()

# ── Build workbook ─────────────────────────────────────────────────────────────

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    return cell


def make_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 120

    rows = [
        ("ANNOTATION INSTRUCTIONS", HEADER_FILL, HEADER_FONT),
        ("", None, None),
        ("PURPOSE", SUBHEAD_FILL, SUBHEAD_FONT),
        ("You are one of two independent annotators validating the automated claim extraction and change classification", None, BOLD_FONT),
        ("produced by an LLM (GPT-4o mini) for a scientific study on how claims change from preprint to publication.", None, NORMAL_FONT),
        ("", None, None),
        ("YOUR TASK", SUBHEAD_FILL, SUBHEAD_FONT),
        ("Go to the 'Annotation' sheet. For each of the 200 rows, fill in TWO columns:", None, BOLD_FONT),
        ("", None, None),
        ("  Column Q  (Rater_1_Extraction_OK)  — if YOU are Rater 1", None, NORMAL_FONT),
        ("  Column R  (Rater_1_Change_OK)       — if YOU are Rater 1", None, NORMAL_FONT),
        ("  Column S  (Rater_2_Extraction_OK)  — if YOU are Rater 2", None, NORMAL_FONT),
        ("  Column T  (Rater_2_Change_OK)       — if YOU are Rater 2", None, NORMAL_FONT),
        ("", None, None),
        ("COLUMN DEFINITIONS", SUBHEAD_FILL, SUBHEAD_FONT),
        ("Extraction_OK: Do the extracted claims (columns G and H) accurately represent what the abstracts say?", None, NORMAL_FONT),
        ("  Enter: YES  if the claims are accurate and complete", None, NORMAL_FONT),
        ("         NO   if one or more claims are wrong, hallucinated, or missing a key point", None, NORMAL_FONT),
        ("         PARTIAL  if most claims are correct but one is slightly off", None, NORMAL_FONT),
        ("", None, None),
        ("Change_OK: Is the LLM's dominant change classification (column J) correct?", None, NORMAL_FONT),
        ("  Enter: YES  if you agree with the classification", None, NORMAL_FONT),
        ("         NO   if you would classify it differently (write your label in the Notes column)", None, NORMAL_FONT),
        ("         UNSURE  if the abstracts are too similar to judge", None, NORMAL_FONT),
        ("", None, None),
        ("IMPORTANT RULES", SUBHEAD_FILL, SUBHEAD_FONT),
        ("1. Do NOT discuss your annotations with the other rater until both have finished.", None, BOLD_FONT),
        ("2. Base your judgment ONLY on the abstract text shown — do not look up the paper.", None, NORMAL_FONT),
        ("3. If an abstract is missing or too short to judge, enter SKIP in both columns.", None, NORMAL_FONT),
        ("4. Use the Notes column (column U or V) freely.", None, NORMAL_FONT),
        ("", None, None),
        ("CHANGE TYPE DEFINITIONS (see also the Codebook sheet)", SUBHEAD_FILL, SUBHEAD_FONT),
        ("strengthened : The published version expresses MORE certainty, stronger evidence, or bolder conclusions", None, NORMAL_FONT),
        ("weakened     : The published version expresses LESS certainty, more hedging, or softer conclusions", None, NORMAL_FONT),
        ("unchanged    : The claim is essentially the same in both versions", None, NORMAL_FONT),
        ("removed      : A claim present in the preprint is absent in the publication", None, NORMAL_FONT),
        ("added        : A claim is present in the publication but was not in the preprint", None, NORMAL_FONT),
    ]

    for i, (text, fill, font) in enumerate(rows, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if fill:  cell.fill  = fill
        if font:  cell.font  = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"


def make_codebook_sheet(wb):
    ws = wb.create_sheet("Codebook")
    headers = ["Change Type", "Definition", "Example (preprint → publication)"]
    widths  = [20, 60, 80]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        set_cell(ws, 1, col, h, font=HEADER_FONT, fill=HEADER_FILL,
                 alignment=Alignment(wrap_text=True, vertical="center"))
        ws.column_dimensions[get_column_letter(col)].width = w

    data = [
        ("strengthened",
         "The published claim expresses greater certainty, stronger causal language, "
         "or more definitive conclusions than the preprint version.",
         "Preprint: 'Our results suggest X may improve Y.'\n"
         "Published: 'Our results demonstrate that X significantly improves Y.'"),
        ("weakened",
         "The published claim expresses less certainty, more hedging, or softer "
         "conclusions than the preprint version.",
         "Preprint: 'X causes Y in all tested conditions.'\n"
         "Published: 'X appears to be associated with Y under certain conditions.'"),
        ("unchanged",
         "The claim conveys essentially the same meaning in both versions, "
         "even if the wording differs slightly.",
         "Preprint: 'Model A outperforms Model B on dataset C.'\n"
         "Published: 'Model A achieves higher accuracy than Model B on dataset C.'"),
        ("removed",
         "A claim present in the preprint abstract is absent from the "
         "published abstract.",
         "Preprint: 'We also show that X generalises to domain D.'\n"
         "Published: [no mention of domain D]"),
        ("added",
         "A claim appears in the published abstract that was not present "
         "in the preprint abstract.",
         "Preprint: [no mention of limitation]\n"
         "Published: 'We note that our approach is limited to English-language texts.'"),
    ]

    fills = [GREEN_FILL, YELLOW_FILL, GREY_FILL, WHITE_FILL, GREEN_FILL]
    for row_i, (ct, defn, example) in enumerate(data, 2):
        fill = fills[row_i % len(fills)]
        for col_i, val in enumerate([ct, defn, example], 1):
            set_cell(ws, row_i, col_i, val, font=NORMAL_FONT, fill=fill,
                     alignment=Alignment(wrap_text=True, vertical="top"),
                     border=THIN_BORDER)
        ws.row_dimensions[row_i].height = 80


def make_annotation_sheet(wb, sampled, corpus_df):
    ws = wb.create_sheet("Annotation")

    columns = [
        ("Row",                    8),
        ("Pair_ID",               20),
        ("Source",                10),
        ("Preprint_Year",         14),
        ("Venue_Type",            18),
        ("Dominant_Change_LLM",   22),
        ("Preprint_Abstract",     60),
        ("Publication_Abstract",  60),
        ("Preprint_Claims_LLM",   55),
        ("Publication_Claims_LLM",55),
        ("Comparisons_LLM",       70),
        ("Pub_DOI",               35),
        ("Preprint_DOI",          35),
        # Rater columns
        ("Rater_1_Extraction_OK", 22),
        ("Rater_1_Change_OK",     20),
        ("Rater_2_Extraction_OK", 22),
        ("Rater_2_Change_OK",     20),
        ("Rater_1_Notes",         35),
        ("Rater_2_Notes",         35),
    ]

    # Header row
    for col_i, (name, width) in enumerate(columns, 1):
        cell = set_cell(ws, 1, col_i, name,
                        font=HEADER_FONT, fill=HEADER_FILL,
                        alignment=Alignment(wrap_text=True,
                                            horizontal="center",
                                            vertical="center"),
                        border=THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_i)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    rater_cols = {14, 15, 16, 17, 18, 19}  # columns for rater input

    for row_i, rec in enumerate(sampled, 2):
        pre_doi = rec.get("preprint_doi", "")
        pub_doi = rec.get("pub_doi", "")

        # Look up abstracts from corpus
        pre_abstract = ""
        pub_abstract = ""
        if pre_doi and pre_doi in corpus_df.index:
            pre_abstract = str(corpus_df.loc[pre_doi, "preprint_abstract"]) \
                if "preprint_abstract" in corpus_df.columns else ""
            pub_abstract = str(corpus_df.loc[pre_doi, "pub_abstract"]) \
                if "pub_abstract" in corpus_df.columns else ""

        pre_claims = fmt_claims(rec.get("preprint_claims", []))
        pub_claims = fmt_claims(rec.get("pub_claims",      []))
        comparisons = fmt_comparisons(rec.get("comparisons", []))

        row_data = [
            row_i - 1,
            rec.get("pair_id", f"pair_{row_i-1}"),
            rec.get("source", ""),
            rec.get("preprint_year", ""),
            rec.get("venue_type", ""),
            rec.get("dominant_change", ""),
            pre_abstract,
            pub_abstract,
            pre_claims,
            pub_claims,
            comparisons,
            pub_doi,
            pre_doi,
            "",  # Rater 1 extraction
            "",  # Rater 1 change
            "",  # Rater 2 extraction
            "",  # Rater 2 change
            "",  # Rater 1 notes
            "",  # Rater 2 notes
        ]

        fill = GREY_FILL if row_i % 2 == 0 else WHITE_FILL

        for col_i, val in enumerate(row_data, 1):
            is_rater = col_i in rater_cols
            cell_fill = YELLOW_FILL if is_rater else fill
            set_cell(ws, row_i, col_i, val,
                     font=NORMAL_FONT,
                     fill=cell_fill,
                     alignment=Alignment(wrap_text=True, vertical="top"),
                     border=THIN_BORDER)

        ws.row_dimensions[row_i].height = 120

    # Add data validation hint in header for rater columns
    ws.cell(row=1, column=14).comment = None  # openpyxl comments optional


def main():
    print("=== Step 11b: Generating Annotation Spreadsheet ===\n")

    print("Loading claim changes...")
    records = load_changes()
    print(f"  Loaded {len(records):,} valid comparison records")

    print("Loading corpus...")
    corpus_df = load_corpus()
    print(f"  Loaded {len(corpus_df):,} corpus rows")

    print(f"\nSampling {SAMPLE_N} pairs (stratified by dominant change type)...")
    sampled = sample_records(records, n=SAMPLE_N)
    print(f"  Sampled {len(sampled)} pairs")

    # Show breakdown
    from collections import Counter
    dc_counts = Counter(r.get("dominant_change", "other") for r in sampled)
    for dc, cnt in sorted(dc_counts.items()):
        print(f"    {dc:<15} : {cnt}")

    print("\nBuilding Excel workbook...")
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    make_instructions_sheet(wb)
    make_annotation_sheet(wb, sampled, corpus_df)
    make_codebook_sheet(wb)

    wb.save(OUT_FILE)
    print(f"\n  Saved → {OUT_FILE}")
    print(f"\nNext steps:")
    print(f"  1. Share annotation_sheet.xlsx with your second annotator")
    print(f"  2. Each annotator fills in their two columns independently")
    print(f"  3. Once both are done, run:")
    print(f"       python3 code/analysis/compute_agreement.py")
    print(f"\nStep 11b complete.")


if __name__ == "__main__":
    main()
